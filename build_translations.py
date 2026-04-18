from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE_FILE = ROOT / "final project 2 2.xlsx"
OUTPUT_FILE = ROOT / "translations-data.js"

REQUIRED_COLUMNS = [
    "French Original",
    "French Clean",
    "English",
    "Arabic",
    "Category",
]

def as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_category(value: object) -> str:
    text = as_text(value)
    if not text:
        return ""

    # Remove exported numeric prefixes such as "107 Criminel [Crim.]".
    text = re.sub(r"^\s*[\d.]+\s*", "", text)
    return text.strip()


def export_translations() -> int:
    if not SOURCE_FILE.exists():
        print(f"خطأ: ملف المصدر غير موجود في {SOURCE_FILE}")
        return 0

    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header = next(worksheet.iter_rows(values_only=True))
    columns = {name: index for index, name in enumerate(header)}

    for col_name in REQUIRED_COLUMNS:
        if col_name not in columns:
            raise ValueError(f"العمود المطلوب '{col_name}' غير موجود في ملف Excel.")

    records: list[dict[str, str]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record = {
            "original": as_text(row[columns["French Original"]]),
            "clean": as_text(row[columns["French Clean"]]),
            "english": as_text(row[columns["English"]]),
            "arabic": as_text(row[columns["Arabic"]]),
            "category": clean_category(row[columns["Category"]]),
        }

        if any(record.values()):
            records.append(record)

    payload = "window.TRANSLATIONS = " + json.dumps(
        records,
        ensure_ascii=False,
        separators=(",", ":"),
    ) + ";\n"

    OUTPUT_FILE.write_text(payload, encoding="utf-8")
    return len(records)


if __name__ == "__main__":
    total = export_translations()
    print(f"Exported {total} translations to {OUTPUT_FILE.name}")
